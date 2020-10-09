# -*- coding: utf-8 -*-
"""
Created on Wed Oct  7 09:15:40 2020

@author: Adewole
"""

import matplotlib.pyplot as plt
import pandas as pd
import openpyxl

# set directory

df = pd.read_excel('2014.xlsx', 'Feb')
wb = openpyxl.load_workbook('2014.xlsx')
#print(wb.sheetnames)
sheet = wb['Aug']

a=sheet['C1']

#print(sheet.cell(row=1,column=2))
#for i in range(1,24,1): print(i,sheet.cell(row=i, column=3).value)
for i in range(1,24): print(i,sheet.cell(row=i, column=4).value)
# set plot

a = df['Time'].astype(str)
b = df['Temperature'].astype(str)
c= df['Dew Point'].astype(str)
d = df['Humidity'].astype(str)
e= df['Wind'].astype(str)
f = df['Wind Speed'].astype(str)
g = df['Wind Gust'].astype(str)
h = df['Pressure'].astype(str)
i = df['Precip'].astype(str)
j = df['Condition'].astype(str)

print(a)
print(b)
#print(c)
#print(d)
#print(e)
#print(f)
#print(g)
#print(h)
#print(i)
#print(j)

plt.plot(a,b)


